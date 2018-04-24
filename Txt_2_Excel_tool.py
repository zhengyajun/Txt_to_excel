import openpyxl
import re
#import json
import os

def readFile():
    author_list = []
    book_list = []
    
    with open("书单.txt") as f:
        for line in f:
        #line = f.readline()
            prog = re.compile(".*\-.*")  
            res = prog.match(line)          # 读取作者名
            if res != None:
                author = line.split()
                author_list.append(author)
           
            else:
                if line[0] == '《':          # 读取书名
                    book = [line.rstrip(), author[1]+author[0]]
                    book_list.append(book)
                    
        write2Excel(book_list, sheet_name="book")
        write2Excel(author_list, sheet_name="author")
        #print(author_list)
        #print(book_list)


def write2Excel(value, path='.\\书单.xlsx', sheet_name=""):
    # wb = openpyxl.load_workbook(path)         # 读取
    # print(wb.sheetnames)
    flag = os.path.exists(path)                 # 判断文件是否存在，已存在执行if，否则else
    if (flag==True):
        wb = openpyxl.load_workbook(path)
        sheet_list = wb.sheetnames
        if (sheet_name not in sheet_list):          # 判断表是否已经存在
            sheet = wb.create_sheet()               # 新加一个表
            sheet.title = sheet_name
            
            for i in range(0, len(value)):
                for j in range(0, len(value[i])):
                    sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
            wb.save(path)           # 保存至path
            print("写入数据成功！")
            #print(wb.sheetnames)   # 显示已有的sheet，测试用
            
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active               # 创建一个工作表    
        sheet.title = sheet_name        # 更改表名     
       
        for i in range(0, len(value)):
            for j in range(0, len(value[i])):
                sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
        wb.save(path)     # 保存至path
        print("写入数据成功！")


readFile()